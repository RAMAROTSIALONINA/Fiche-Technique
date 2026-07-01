from fastapi import FastAPI, HTTPException, Depends, File, UploadFile, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from starlette.middleware.base import BaseHTTPMiddleware
from sqlalchemy import create_engine, Column, String, Integer, Float, Text, ForeignKey, text
from sqlalchemy.orm import declarative_base, sessionmaker, Session, relationship
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime
import random, string, os, sys, uvicorn, re, json, hashlib, secrets

try:
    import fitz  # pymupdf
    HAS_FITZ = True
except ImportError:
    HAS_FITZ = False

import io

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── APP ───────────────────────────────────────────────────────────────────────
app = FastAPI(title="Fiche Technique BOGOTA")

# ── DATABASE ──────────────────────────────────────────────────────────────────
_FROZEN = getattr(sys, 'frozen', False)

if _FROZEN:
    BASE_DIR   = os.path.dirname(sys.executable)
    STATIC_DIR = os.path.join(sys._MEIPASS, 'static')
    # Base de données sur D: si le disque existe et que l'app n'est pas déjà sur D:
    _D_DATA = r"D:\FicheTechnique-BOGOTA"
    if os.path.exists("D:\\") and not BASE_DIR.upper().startswith("D:"):
        DATA_DIR = _D_DATA
        os.makedirs(DATA_DIR, exist_ok=True)
    else:
        DATA_DIR = BASE_DIR   # App sur D: ou pas de D: → données à côté du .exe
else:
    BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
    STATIC_DIR = os.path.join(BASE_DIR, 'static')
    DATA_DIR   = BASE_DIR

DATABASE_URL = f"sqlite:///{os.path.join(DATA_DIR, 'fiches_bogota.db')}"

engine       = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
Base         = declarative_base()

# ── MODELS ────────────────────────────────────────────────────────────────────
class Categorie(Base):
    __tablename__ = "categories"
    id      = Column(String, primary_key=True)
    nom     = Column(String, nullable=False)
    couleur = Column(String, default="#8A8480")
    groupe  = Column(String, default="PRODUCTION")  # PREPA ou PRODUCTION

class Allergene(Base):
    __tablename__ = "allergenes"
    nom = Column(String, primary_key=True)

class Ingredient(Base):
    __tablename__ = "ingredients"
    id       = Column(Integer, primary_key=True, autoincrement=True)
    fiche_id = Column(String, ForeignKey("fiches.id", ondelete="CASCADE"))
    nom      = Column(String)
    qte      = Column(String)
    unite    = Column(String)
    prix     = Column(Float)

class Etape(Base):
    __tablename__ = "etapes"
    id       = Column(Integer, primary_key=True, autoincrement=True)
    fiche_id = Column(String, ForeignKey("fiches.id", ondelete="CASCADE"))
    ordre    = Column(Integer)
    texte    = Column(Text)

class FicheAllergene(Base):
    __tablename__ = "fiche_allergenes"
    fiche_id  = Column(String, ForeignKey("fiches.id", ondelete="CASCADE"), primary_key=True)
    allergene = Column(String, primary_key=True)

class Fiche(Base):
    __tablename__ = "fiches"
    id                 = Column(String, primary_key=True)
    nom                = Column(String, nullable=False)
    categorie          = Column(String, ForeignKey("categories.id"))
    quantite           = Column(Integer)
    unite              = Column(String, default="personnes")
    cout               = Column(Float)
    coefficient        = Column(Float, default=2.0)
    conservation_val   = Column(Integer)
    conservation_unite = Column(String, default="jours")
    prep               = Column(Integer)
    cuisson            = Column(Integer)
    temperature        = Column(Integer)
    description        = Column(Text)
    photo              = Column(Text)
    apparence_normale  = Column(Text)
    apparence_anormale = Column(Text)
    materiels          = Column(Text)   # séparé par \n
    utilisations       = Column(Text)
    notes_preparateur  = Column(Text)
    notes_services     = Column(Text)
    created_at         = Column(String)
    updated_at         = Column(String)

    ingredients = relationship("Ingredient", cascade="all, delete-orphan", order_by="Ingredient.id")
    etapes      = relationship("Etape",      cascade="all, delete-orphan", order_by="Etape.ordre")
    allergenes  = relationship("FicheAllergene", cascade="all, delete-orphan")

class Historique(Base):
    __tablename__ = "historique"
    id         = Column(Integer, primary_key=True, autoincrement=True)
    action     = Column(String)   # créer / modifier / supprimer
    fiche_id   = Column(String)
    fiche_nom  = Column(String)
    horodatage = Column(String)   # ISO datetime UTC
    details    = Column(Text)     # JSON : changements, infos

class User(Base):
    __tablename__ = "users"
    id            = Column(Integer, primary_key=True, autoincrement=True)
    username      = Column(String, unique=True, nullable=False)
    password_hash = Column(String, nullable=False)
    role          = Column(String, default="user")   # admin / user
    is_active     = Column(Integer, default=1)
    created_at    = Column(String)

class UserSession(Base):
    __tablename__ = "user_sessions"
    token      = Column(String, primary_key=True)
    user_id    = Column(Integer, ForeignKey("users.id", ondelete="CASCADE"))
    created_at = Column(String)

class IngredientCatalogue(Base):
    __tablename__ = "ingredients_catalogue"
    id    = Column(Integer, primary_key=True, autoincrement=True)
    nom   = Column(String, nullable=False, unique=True)
    unite = Column(String)
    prix  = Column(Float)

Base.metadata.create_all(bind=engine)

# ── MIGRATION ─────────────────────────────────────────────────────────────────
def migrate():
    """Ajoute les nouvelles colonnes si elles n'existent pas encore."""
    with engine.connect() as conn:
        result  = conn.execute(text("PRAGMA table_info(fiches)"))
        existing = {row[1] for row in result}
        new_cols = {
            'apparence_normale':  'TEXT',
            'apparence_anormale': 'TEXT',
            'materiels':          'TEXT',
            'utilisations':       'TEXT',
            'notes_preparateur':  'TEXT',
            'notes_services':     'TEXT',
        }
        for col, typ in new_cols.items():
            if col not in existing:
                conn.execute(text(f'ALTER TABLE fiches ADD COLUMN {col} {typ}'))
        # Migration table categories — colonne groupe
        cat_info = conn.execute(text("PRAGMA table_info(categories)"))
        cat_cols = {row[1] for row in cat_info}
        if 'groupe' not in cat_cols:
            conn.execute(text("ALTER TABLE categories ADD COLUMN groupe TEXT DEFAULT 'PRODUCTION'"))
            conn.execute(text("UPDATE categories SET groupe='PREPA' WHERE id IN ('prepa_cuisine','prepa_sauce','sauces')"))
        # Migration table historique
        hist_info = conn.execute(text("PRAGMA table_info(historique)"))
        hist_cols = {row[1] for row in hist_info}
        if 'details' not in hist_cols:
            conn.execute(text('ALTER TABLE historique ADD COLUMN details TEXT'))
        # Migration table ingredients — colonne prix
        ing_info = conn.execute(text("PRAGMA table_info(ingredients)"))
        ing_cols = {row[1] for row in ing_info}
        if 'prix' not in ing_cols:
            conn.execute(text('ALTER TABLE ingredients ADD COLUMN prix REAL'))
        conn.commit()

migrate()

# ── PYDANTIC SCHEMAS ──────────────────────────────────────────────────────────
class IngredientIn(BaseModel):
    nom: str; qte: str; unite: str; prix: Optional[float] = None

class IngredientCatalogueIn(BaseModel):
    nom: str
    unite: Optional[str] = None
    prix: Optional[float] = None

class FicheIn(BaseModel):
    id:                 Optional[str]   = None
    nom:                str
    categorie:          Optional[str]   = None
    quantite:           Optional[int]   = None
    unite:              Optional[str]   = "personnes"
    cout:               Optional[float] = None
    coefficient:        Optional[float] = 2.0
    conservation_val:   Optional[int]   = None
    conservation_unite: Optional[str]   = "jours"
    prep:               Optional[int]   = None
    cuisson:            Optional[int]   = None
    temperature:        Optional[int]   = None
    description:        Optional[str]       = None
    photo:              Optional[str]       = None
    ingredients:        Optional[List[IngredientIn]] = []
    etapes:             Optional[List[str]] = []
    allergenes:         Optional[List[str]] = []
    apparence_normale:  Optional[str]       = None
    apparence_anormale: Optional[str]       = None
    materiels:          Optional[List[str]] = []
    utilisations:       Optional[str]       = None
    notes_preparateur:  Optional[str]       = None
    notes_services:     Optional[str]       = None

class CategorieIn(BaseModel):
    id:      Optional[str] = None
    nom:     str
    couleur: Optional[str] = "#8A8480"
    groupe:  Optional[str] = "PRODUCTION"

class AllergeneIn(BaseModel):
    nom: str

class LoginIn(BaseModel):
    username: str
    password: str

class UserIn(BaseModel):
    username: str
    password: Optional[str] = None
    role:     Optional[str] = "user"
    is_active:Optional[int] = 1

class ChangePasswordIn(BaseModel):
    old_password: str
    new_password: str

# ── HELPERS ───────────────────────────────────────────────────────────────────
def get_db():
    db = SessionLocal()
    try: yield db
    finally: db.close()

# ── AUTH HELPERS ──────────────────────────────────────────────────────────────
def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode("utf-8")).hexdigest()

def verify_password(password: str, password_hash: str) -> bool:
    return hash_password(password) == password_hash

def create_default_admin():
    db = SessionLocal()
    try:
        if db.query(User).count() == 0:
            db.add(User(
                username="admin",
                password_hash=hash_password("admin123"),
                role="admin",
                is_active=1,
                created_at=datetime.utcnow().isoformat()
            ))
            db.commit()
            print("Compte admin par defaut cree (admin / admin123)")
    finally:
        db.close()

# ── MIDDLEWARE AUTH ────────────────────────────────────────────────────────────
PUBLIC_PATHS = {"/api/auth/login"}

class AuthMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        path = request.url.path
        # Autorisé sans token : login, fichiers statiques, page principale
        if path in PUBLIC_PATHS or not path.startswith("/api/"):
            return await call_next(request)
        token = request.headers.get("Authorization", "").replace("Bearer ", "").strip()
        if not token:
            return JSONResponse({"detail": "Non authentifie"}, status_code=401)
        db = SessionLocal()
        try:
            sess = db.query(UserSession).filter(UserSession.token == token).first()
            if not sess:
                return JSONResponse({"detail": "Session invalide"}, status_code=401)
            user = db.query(User).filter(User.id == sess.user_id, User.is_active == 1).first()
            if not user:
                return JSONResponse({"detail": "Utilisateur introuvable"}, status_code=401)
            request.state.current_user = user
        finally:
            db.close()
        return await call_next(request)

app.add_middleware(AuthMiddleware)
create_default_admin()

def get_current_user(request: Request, db: Session = Depends(get_db)):
    return getattr(request.state, "current_user", None)

def gen_id(prefix=""):
    return prefix + datetime.utcnow().strftime("%Y%m%d%H%M%S") + \
           ''.join(random.choices(string.ascii_lowercase, k=4))

def fiche_to_dict(f):
    return {
        "id": f.id, "nom": f.nom, "categorie": f.categorie,
        "quantite": f.quantite, "unite": f.unite,
        "cout": f.cout, "coefficient": f.coefficient,
        "conservation_val": f.conservation_val, "conservation_unite": f.conservation_unite,
        "prep": f.prep, "cuisson": f.cuisson, "temperature": f.temperature,
        "description": f.description, "photo": f.photo,
        "ingredients": [{"nom": i.nom, "qte": i.qte, "unite": i.unite, "prix": i.prix} for i in f.ingredients],
        "etapes":      [e.texte for e in sorted(f.etapes, key=lambda x: x.ordre)],
        "allergenes":  [a.allergene for a in f.allergenes],
        "apparence_normale":  f.apparence_normale,
        "apparence_anormale": f.apparence_anormale,
        "materiels":   [m.strip() for m in (f.materiels or '').split('\n') if m.strip()],
        "utilisations":       f.utilisations,
        "notes_preparateur":  f.notes_preparateur,
        "notes_services":     f.notes_services,
        "createdAt": f.created_at, "updatedAt": f.updated_at,
    }

def save_relations(db, fiche_id, ingredients, etapes, allergenes):
    db.query(Ingredient).filter(Ingredient.fiche_id == fiche_id).delete()
    db.query(Etape).filter(Etape.fiche_id == fiche_id).delete()
    db.query(FicheAllergene).filter(FicheAllergene.fiche_id == fiche_id).delete()
    for ing in (ingredients or []):
        if ing.nom.strip():
            db.add(Ingredient(fiche_id=fiche_id, nom=ing.nom, qte=ing.qte, unite=ing.unite, prix=ing.prix))
    for i, etape in enumerate(etapes or []):
        if etape.strip():
            db.add(Etape(fiche_id=fiche_id, ordre=i+1, texte=etape))
    for allergen in (allergenes or []):
        db.add(FicheAllergene(fiche_id=fiche_id, allergene=allergen))

def log_history(db, action: str, fiche_id: str, fiche_nom: str, details: dict = None, par: str = None, type_: str = "fiche"):
    d = {"type": type_}
    if par:   d["par"] = par
    if details: d.update(details)
    db.add(Historique(action=action, fiche_id=str(fiche_id), fiche_nom=fiche_nom,
                      horodatage=datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
                      details=json.dumps(d, ensure_ascii=False)))

def require_admin(request: Request, db=None):
    user = getattr(request.state, "current_user", None)
    if not user or user.role != "admin":
        raise HTTPException(403, "Accès réservé aux administrateurs")

# ── ROUTES : AUTHENTIFICATION ────────────────────────────────────────────────
@app.post("/api/auth/login")
def login(data: LoginIn, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == data.username, User.is_active == 1).first()
    if not user or not verify_password(data.password, user.password_hash):
        raise HTTPException(status_code=401, detail="Identifiants incorrects")
    token = secrets.token_hex(32)
    db.add(UserSession(token=token, user_id=user.id, created_at=datetime.utcnow().isoformat()))
    log_history(db, "connexion", str(user.id), user.username, {"role": user.role}, par=user.username, type_="auth")
    db.commit()
    return {"token": token, "user": {"id": user.id, "username": user.username, "role": user.role}}

@app.post("/api/auth/logout")
def logout(request: Request, db: Session = Depends(get_db)):
    token = request.headers.get("Authorization", "").replace("Bearer ", "").strip()
    if token:
        db.query(UserSession).filter(UserSession.token == token).delete()
        db.commit()
    return {"ok": True}

@app.get("/api/auth/me")
def me(request: Request):
    user = getattr(request.state, "current_user", None)
    if not user:
        raise HTTPException(status_code=401, detail="Non authentifie")
    return {"id": user.id, "username": user.username, "role": user.role}

@app.put("/api/auth/change-password")
def change_password(data: ChangePasswordIn, request: Request, db: Session = Depends(get_db)):
    user = getattr(request.state, "current_user", None)
    if not user or not verify_password(data.old_password, user.password_hash):
        raise HTTPException(status_code=400, detail="Ancien mot de passe incorrect")
    if len(data.new_password) < 4:
        raise HTTPException(status_code=400, detail="Mot de passe trop court (min 4 caracteres)")
    db_user = db.query(User).filter(User.id == user.id).first()
    db_user.password_hash = hash_password(data.new_password)
    log_history(db, "modifier", str(user.id), user.username, {"champ": "mot de passe"}, par=user.username, type_="auth")
    db.commit()
    return {"ok": True}

# ── ROUTES : GESTION UTILISATEURS (admin) ────────────────────────────────────
@app.get("/api/users")
def list_users(request: Request, db: Session = Depends(get_db)):
    require_admin(request, db)
    users = db.query(User).order_by(User.id).all()
    return [{"id": u.id, "username": u.username, "role": u.role,
             "is_active": u.is_active, "created_at": u.created_at} for u in users]

@app.post("/api/users", status_code=201)
def create_user(data: UserIn, request: Request, db: Session = Depends(get_db)):
    require_admin(request, db)
    if not data.password or len(data.password) < 4:
        raise HTTPException(status_code=400, detail="Mot de passe trop court (min 4 caracteres)")
    if db.query(User).filter(User.username == data.username).first():
        raise HTTPException(status_code=400, detail="Nom d'utilisateur deja utilise")
    if data.role not in ("admin", "user"):
        raise HTTPException(status_code=400, detail="Role invalide (admin ou user)")
    actor = getattr(request.state, "current_user", None)
    u = User(username=data.username, password_hash=hash_password(data.password),
             role=data.role, is_active=data.is_active, created_at=datetime.utcnow().isoformat())
    db.add(u); db.flush()
    log_history(db, "créer", str(u.id), u.username, {"role": u.role}, par=actor.username if actor else None, type_="utilisateur")
    db.commit(); db.refresh(u)
    return {"id": u.id, "username": u.username, "role": u.role, "is_active": u.is_active}

@app.put("/api/users/{user_id}")
def update_user(user_id: int, data: UserIn, request: Request, db: Session = Depends(get_db)):
    require_admin(request, db)
    current = getattr(request.state, "current_user", None)
    u = db.query(User).filter(User.id == user_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="Utilisateur introuvable")
    if u.username != data.username and db.query(User).filter(User.username == data.username).first():
        raise HTTPException(status_code=400, detail="Nom d'utilisateur deja utilise")
    if data.role not in ("admin", "user"):
        raise HTTPException(status_code=400, detail="Role invalide")
    # Empêche de désactiver/dégrader le seul admin
    if current and u.id == current.id and data.role != "admin":
        raise HTTPException(status_code=400, detail="Vous ne pouvez pas changer votre propre role")
    changes = []
    if u.username  != data.username:  changes.append({"champ":"nom",    "avant": u.username,       "apres": data.username})
    if u.role      != data.role:      changes.append({"champ":"role",   "avant": u.role,           "apres": data.role})
    if u.is_active != data.is_active: changes.append({"champ":"statut", "avant": str(u.is_active), "apres": str(data.is_active)})
    if data.password:                 changes.append({"champ":"mot de passe", "avant": "***", "apres": "***"})
    u.username  = data.username
    u.role      = data.role
    u.is_active = data.is_active
    if data.password:
        if len(data.password) < 4:
            raise HTTPException(status_code=400, detail="Mot de passe trop court (min 4 caracteres)")
        u.password_hash = hash_password(data.password)
    log_history(db, "modifier", str(u.id), u.username, {"changes": changes}, par=current.username if current else None, type_="utilisateur")
    db.commit()
    return {"id": u.id, "username": u.username, "role": u.role, "is_active": u.is_active}

@app.delete("/api/users/{user_id}")
def delete_user(user_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin(request, db)
    current = getattr(request.state, "current_user", None)
    if current and current.id == user_id:
        raise HTTPException(status_code=400, detail="Vous ne pouvez pas supprimer votre propre compte")
    u = db.query(User).filter(User.id == user_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="Utilisateur introuvable")
    actor = getattr(request.state, "current_user", None)
    log_history(db, "supprimer", str(u.id), u.username, {"role": u.role}, par=actor.username if actor else None, type_="utilisateur")
    db.query(UserSession).filter(UserSession.user_id == user_id).delete()
    db.delete(u); db.commit()
    return {"ok": True}

# ── ROUTES : FICHES ───────────────────────────────────────────────────────────
@app.get("/api/fiches")
def list_fiches(db: Session = Depends(get_db)):
    return [fiche_to_dict(f) for f in
            db.query(Fiche).order_by(Fiche.updated_at.desc()).all()]

@app.post("/api/fiches", status_code=201)
def create_fiche(data: FicheIn, request: Request, db: Session = Depends(get_db)):
    now = datetime.utcnow().isoformat()
    f   = Fiche(
        id=data.id or gen_id(), nom=data.nom, categorie=data.categorie,
        quantite=data.quantite, unite=data.unite, cout=data.cout,
        coefficient=data.coefficient, conservation_val=data.conservation_val,
        conservation_unite=data.conservation_unite, prep=data.prep,
        cuisson=data.cuisson, temperature=data.temperature,
        description=data.description, photo=data.photo,
        apparence_normale=data.apparence_normale,
        apparence_anormale=data.apparence_anormale,
        materiels='\n'.join(data.materiels) if data.materiels else None,
        utilisations=data.utilisations,
        notes_preparateur=data.notes_preparateur,
        notes_services=data.notes_services,
        created_at=now, updated_at=now
    )
    actor = getattr(request.state, "current_user", None)
    db.add(f); db.flush()
    save_relations(db, f.id, data.ingredients, data.etapes, data.allergenes)
    log_history(db, "créer", f.id, f.nom, {
        "categorie": data.categorie,
        "cout": data.cout,
        "nb_ingredients": len(data.ingredients or []),
        "nb_etapes": len(data.etapes or []),
    }, par=actor.username if actor else None)
    db.commit(); db.refresh(f)
    return fiche_to_dict(f)

@app.get("/api/fiches/{fiche_id}")
def get_fiche(fiche_id: str, db: Session = Depends(get_db)):
    f = db.query(Fiche).filter(Fiche.id == fiche_id).first()
    if not f: raise HTTPException(404, "Fiche non trouvée")
    return fiche_to_dict(f)

@app.put("/api/fiches/{fiche_id}")
def update_fiche(fiche_id: str, data: FicheIn, request: Request, db: Session = Depends(get_db)):
    f = db.query(Fiche).filter(Fiche.id == fiche_id).first()
    if not f: raise HTTPException(404, "Fiche non trouvée")
    # Capturer les changements avant modification
    def _fmt(v, lbl):
        if v is None: return "—"
        if lbl in ("Photo",) : return "✔" if v else "—"
        if lbl in ("Description","App. normale","App. anormale","Utilisations","Notes prép.","Notes service"):
            s = str(v).strip()
            return (s[:40] + "…") if len(s) > 40 else s
        return str(v)

    _checks = [
        ("Nom",           f.nom,              data.nom),
        ("Catégorie",     f.categorie,         data.categorie),
        ("Coût (Ar)",     f.cout,              data.cout),
        ("Coefficient",   f.coefficient,       data.coefficient),
        ("Quantité",      f.quantite,          data.quantite),
        ("Unité",         f.unite,             data.unite),
        ("Préparation",   f.prep,              data.prep),
        ("Cuisson",       f.cuisson,           data.cuisson),
        ("Température",   f.temperature,       data.temperature),
        ("Conservation",  f.conservation_val,  data.conservation_val),
        ("Photo",         f.photo,             data.photo),
        ("Description",   f.description,       data.description),
        ("App. normale",  f.apparence_normale, data.apparence_normale),
        ("App. anormale", f.apparence_anormale,data.apparence_anormale),
        ("Utilisations",  f.utilisations,      data.utilisations),
        ("Notes prép.",   f.notes_preparateur, data.notes_preparateur),
        ("Notes service", f.notes_services,    data.notes_services),
    ]
    changes = [{"champ": lbl, "avant": _fmt(old, lbl), "apres": _fmt(new, lbl)}
               for lbl, old, new in _checks if old != new]
    old_ing = len(f.ingredients)
    new_ing = len(data.ingredients or [])
    if old_ing != new_ing:
        changes.append({"champ": "Ingrédients", "avant": f"{old_ing}", "apres": f"{new_ing}"})
    for k, v in {"nom": data.nom, "categorie": data.categorie,
                  "quantite": data.quantite, "unite": data.unite,
                  "cout": data.cout, "coefficient": data.coefficient,
                  "conservation_val": data.conservation_val,
                  "conservation_unite": data.conservation_unite,
                  "prep": data.prep, "cuisson": data.cuisson,
                  "temperature": data.temperature, "description": data.description,
                  "photo": data.photo,
                  "apparence_normale": data.apparence_normale,
                  "apparence_anormale": data.apparence_anormale,
                  "materiels": '\n'.join(data.materiels) if data.materiels else None,
                  "utilisations": data.utilisations,
                  "notes_preparateur": data.notes_preparateur,
                  "notes_services": data.notes_services,
                  "updated_at": datetime.utcnow().isoformat()}.items():
        setattr(f, k, v)
    actor = getattr(request.state, "current_user", None)
    save_relations(db, fiche_id, data.ingredients, data.etapes, data.allergenes)
    log_history(db, "modifier", f.id, f.nom, {"changes": changes}, par=actor.username if actor else None)
    db.commit(); db.refresh(f)
    return fiche_to_dict(f)

@app.delete("/api/fiches/{fiche_id}")
def delete_fiche(fiche_id: str, request: Request, db: Session = Depends(get_db)):
    actor = getattr(request.state, "current_user", None)
    f = db.query(Fiche).filter(Fiche.id == fiche_id).first()
    if not f: raise HTTPException(404)
    log_history(db, "supprimer", f.id, f.nom, {
        "categorie": f.categorie,
        "cout": f.cout,
        "nb_ingredients": len(f.ingredients),
        "nb_etapes": len(f.etapes),
    }, par=actor.username if actor else None)
    db.delete(f); db.commit()
    return {"ok": True}

# ── ROUTES : HISTORIQUE ──────────────────────────────────────────────────────
@app.get("/api/historique")
def get_historique(limit: int = 200, db: Session = Depends(get_db)):
    rows = db.query(Historique).order_by(Historique.id.desc()).limit(limit).all()
    return [{"id": h.id, "action": h.action, "fiche_id": h.fiche_id,
             "fiche_nom": h.fiche_nom, "horodatage": h.horodatage,
             "details": json.loads(h.details) if h.details else None} for h in rows]

@app.delete("/api/historique")
def clear_historique(request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    db.query(Historique).delete(); db.commit()
    return {"ok": True}

# ── ROUTES : CATEGORIES ───────────────────────────────────────────────────────
@app.get("/api/categories")
def list_categories(db: Session = Depends(get_db)):
    return [{"id": c.id, "nom": c.nom, "couleur": c.couleur, "groupe": c.groupe or "PRODUCTION"}
            for c in db.query(Categorie).all()]

@app.post("/api/categories", status_code=201)
def create_categorie(data: CategorieIn, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    actor = getattr(request.state, "current_user", None)
    groupe = data.groupe if data.groupe in ("PREPA", "PRODUCTION") else "PRODUCTION"
    c = Categorie(id=data.id or gen_id("cat_"), nom=data.nom, couleur=data.couleur, groupe=groupe)
    db.add(c); db.flush()
    log_history(db, "créer", c.id, c.nom, {"couleur": c.couleur, "groupe": groupe}, par=actor.username if actor else None, type_="categorie")
    db.commit()
    return {"id": c.id, "nom": c.nom, "couleur": c.couleur, "groupe": c.groupe}

@app.put("/api/categories/{cat_id}")
def update_categorie(cat_id: str, data: CategorieIn, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    actor = getattr(request.state, "current_user", None)
    c = db.query(Categorie).filter(Categorie.id == cat_id).first()
    if not c: raise HTTPException(404)
    groupe = data.groupe if data.groupe in ("PREPA", "PRODUCTION") else "PRODUCTION"
    changes = []
    if c.nom    != data.nom:    changes.append({"champ": "nom",    "avant": c.nom,    "apres": data.nom})
    if c.couleur!= data.couleur:changes.append({"champ": "couleur","avant": c.couleur,"apres": data.couleur})
    if c.groupe != groupe:      changes.append({"champ": "groupe", "avant": c.groupe, "apres": groupe})
    c.nom = data.nom; c.couleur = data.couleur; c.groupe = groupe
    log_history(db, "modifier", c.id, c.nom, {"changes": changes}, par=actor.username if actor else None, type_="categorie")
    db.commit()
    return {"id": c.id, "nom": c.nom, "couleur": c.couleur, "groupe": c.groupe}

@app.delete("/api/categories/{cat_id}")
def delete_categorie(cat_id: str, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    actor = getattr(request.state, "current_user", None)
    c = db.query(Categorie).filter(Categorie.id == cat_id).first()
    if not c: raise HTTPException(404)
    log_history(db, "supprimer", c.id, c.nom, {}, par=actor.username if actor else None, type_="categorie")
    db.delete(c); db.commit()
    return {"ok": True}

# ── ROUTE : CATALOGUE INGRÉDIENTS (fiches PREPA uniquement) ──────────────────
@app.get("/api/ingredients/catalogue")
def get_ingredients_catalogue(db: Session = Depends(get_db)):
    prepa_ids = [c.id for c in db.query(Categorie).filter(Categorie.groupe == "PREPA").all()]
    fiches = (
        db.query(Fiche)
        .filter(Fiche.categorie.in_(prepa_ids))
        .order_by(Fiche.nom)
        .all()
    )
    from_prepa = [{"nom": f.nom, "prix": f.cout, "qte": f.quantite, "unite": f.unite} for f in fiches]
    from_manuel = [{"nom": e.nom, "prix": e.prix, "qte": None, "unite": e.unite}
                   for e in db.query(IngredientCatalogue).order_by(IngredientCatalogue.nom).all()]
    # Fusionner sans doublons (prepa prioritaire)
    noms_prepa = {e["nom"] for e in from_prepa}
    merged = from_prepa + [e for e in from_manuel if e["nom"] not in noms_prepa]
    return sorted(merged, key=lambda x: x["nom"].lower())


@app.get("/api/ingredients/catalogue/manuel")
def list_catalogue_manuel(db: Session = Depends(get_db)):
    return [{"id": e.id, "nom": e.nom, "unite": e.unite, "prix": e.prix}
            for e in db.query(IngredientCatalogue).order_by(IngredientCatalogue.nom).all()]

@app.post("/api/ingredients/catalogue/manuel", status_code=201)
def create_catalogue_manuel(data: IngredientCatalogueIn, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    nom = data.nom.strip()
    if not nom: raise HTTPException(400, "Nom vide")
    if db.query(IngredientCatalogue).filter(IngredientCatalogue.nom == nom).first():
        raise HTTPException(409, "Cet ingrédient existe déjà")
    entry = IngredientCatalogue(nom=nom, unite=data.unite, prix=data.prix)
    db.add(entry)
    db.commit()
    db.refresh(entry)
    return {"id": entry.id, "nom": entry.nom, "unite": entry.unite, "prix": entry.prix}

@app.put("/api/ingredients/catalogue/manuel/{entry_id}")
def update_catalogue_manuel(entry_id: int, data: IngredientCatalogueIn, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    entry = db.query(IngredientCatalogue).filter(IngredientCatalogue.id == entry_id).first()
    if not entry: raise HTTPException(404, "Introuvable")
    entry.nom = data.nom.strip()
    entry.unite = data.unite
    entry.prix = data.prix
    db.commit()
    return {"id": entry.id, "nom": entry.nom, "unite": entry.unite, "prix": entry.prix}

@app.delete("/api/ingredients/catalogue/manuel/{entry_id}")
def delete_catalogue_manuel(entry_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    entry = db.query(IngredientCatalogue).filter(IngredientCatalogue.id == entry_id).first()
    if not entry: raise HTTPException(404, "Introuvable")
    db.delete(entry)
    db.commit()
    return {"ok": True}

class RenommerIngredientIn(BaseModel):
    ancien: str
    nouveau: str

class SupprimerIngredientIn(BaseModel):
    nom: str

@app.put("/api/ingredients/fiche/renommer")
def renommer_ingredient_fiche(data: RenommerIngredientIn, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    ancien = data.ancien.strip(); nouveau = data.nouveau.strip()
    if not ancien or not nouveau: raise HTTPException(400, "Nom vide")
    rows = db.query(Ingredient).filter(Ingredient.nom == ancien).all()
    for r in rows: r.nom = nouveau
    db.commit()
    return {"modifie": len(rows)}

@app.post("/api/ingredients/fiche/supprimer")
def supprimer_ingredient_fiche(data: SupprimerIngredientIn, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    nom = data.nom.strip()
    if not nom: raise HTTPException(400, "Nom vide")
    rows = db.query(Ingredient).filter(Ingredient.nom == nom).all()
    count = len(rows)
    for r in rows: db.delete(r)
    db.commit()
    return {"supprime": count}


# ── ROUTES : ALLERGENES ───────────────────────────────────────────────────────
@app.get("/api/allergenes")
def list_allergenes(db: Session = Depends(get_db)):
    return [a.nom for a in db.query(Allergene).all()]

@app.post("/api/allergenes", status_code=201)
def create_allergene(data: AllergeneIn, request: Request, db: Session = Depends(get_db)):
    actor = getattr(request.state, "current_user", None)
    nom = data.nom.strip()
    if not nom: raise HTTPException(400, "Nom vide")
    if not db.query(Allergene).filter(Allergene.nom == nom).first():
        db.add(Allergene(nom=nom))
        log_history(db, "créer", nom, nom, {}, par=actor.username if actor else None, type_="allergene")
        db.commit()
    return {"nom": nom}

@app.delete("/api/allergenes/{nom}")
def delete_allergene(nom: str, request: Request, db: Session = Depends(get_db)):
    actor = getattr(request.state, "current_user", None)
    a = db.query(Allergene).filter(Allergene.nom == nom).first()
    if a:
        log_history(db, "supprimer", nom, nom, {}, par=actor.username if actor else None, type_="allergene")
        db.delete(a); db.commit()
    return {"ok": True}

# ── SEED DATA ─────────────────────────────────────────────────────────────────
def seed():
    db = SessionLocal()
    # (id, nom, couleur, groupe)
    ALL_CATS = [
        # ── PREPA (Semi-fini) ─────────────────────────────────────────────
        ("prepa_cuisine", "Préparations",  "#E67E22", "PREPA"),
        ("prepa_sauce",   "Sauces & Bases","#16A085", "PREPA"),
        # ── PRODUCTION (Produit fini) ──────────────────────────────────────
        ("pizza",   "Pizza",    "#E74C3C", "PRODUCTION"),
        ("boisson", "Boisson",  "#2980B9", "PRODUCTION"),
        ("snack",   "Snack",    "#F39C12", "PRODUCTION"),
        ("autre",   "Autre",    "#95A5A6", "PRODUCTION"),
    ]
    existing = {c.id: c for c in db.query(Categorie).all()}
    for cid, nom, col, grp in ALL_CATS:
        if cid not in existing:
            db.add(Categorie(id=cid, nom=nom, couleur=col, groupe=grp))
        else:
            # Mettre à jour le groupe si manquant
            if not existing[cid].groupe:
                existing[cid].groupe = grp
    if db.query(Allergene).count() == 0:
        for a in ["Gluten","Crustacés","Œufs","Poissons","Arachides","Soja",
                  "Lait","Fruits à coques","Céleri","Moutarde","Sésame",
                  "Anhydride sulfureux","Lupin","Mollusques"]:
            db.add(Allergene(nom=a))
    db.commit(); db.close()

seed()

# ── PARSEUR PDF ───────────────────────────────────────────────────────────────
def parse_fiche_pdf(text: str):
    lines = [l.strip() for l in text.split('\n') if l.strip()]

    fiche = dict(nom=None, categorie=None, cout=None, quantite=None, unite=None,
                 prep=None, etapes=[], ingredients=[], materiels=[],
                 utilisations=None, notes_services=None, notes_preparateur=None,
                 apparence_normale=None, apparence_anormale=None, conservation=None)

    # ── Extractions simples ───────────────────────────────────────────────────
    for line in lines:
        if line.startswith('Section:'):
            fiche['categorie'] = line.replace('Section:', '').strip().lower()
        elif re.match(r'C\.?R\s*:', line, re.I) and 'MGA' in line.upper():
            nums = re.findall(r'\d+', line.split(':')[-1])
            if nums:
                try: fiche['cout'] = float(''.join(nums))
                except: pass
        elif 'Rendement de la recette' in line:
            rend = line.split(':')[-1].strip()
            m = re.search(r'([\d,.]+)\s*([A-Za-zéèàùûîôê]+)?', rend)
            if m:
                try: fiche['quantite'] = int(float(m.group(1).replace(',', '.')))
                except: pass
                fiche['unite'] = (m.group(2) or 'portions').lower()
        elif 'Temps total de Production' in line:
            m = re.search(r'(\d+)', line)
            if m: fiche['prep'] = int(m.group(1))

    # ── Nom de la recette ─────────────────────────────────────────────────────
    # Le nom est toujours juste avant "Notes du Préparateur" dans le PDF
    SKIP_CAPS  = {'NOTES DE SERVICES','APPARENCE NORMALE','APPARENCE ANORMALE',
                  'CONSERVATION','COEF','INGRÉDIENTS','INFORMATION GÉNÉRAL'}
    SKIP_PARTS = ['ÉTAPE','BASE','FOND','MIXAGE','MONTAGE','FINITION','PANURE',
                  'CUISSON','MÉLANGE','DÉGLAÇAGE','REFROIDISSEMENT','PRÉPARATION',
                  'SAUCE ET FINITION','PRÉ-PANURE']

    try:
        np_i = next(i for i,l in enumerate(lines) if l == 'Notes du Préparateur')
        # Chercher dans les 6 lignes précédant "Notes du Préparateur"
        window = lines[max(0, np_i-6): np_i]
        for line in reversed(window):
            if (len(line) >= 2
                    and re.search(r'[A-ZÉÈÀÙÂÊÎÔÛÄËÏÖÜÇ]', line)
                    and not re.match(r'^[\d\s°.,:+×✓→]+$', line)
                    and not re.match(r'^\d', line)
                    and line.strip(' .') not in SKIP_CAPS
                    and not any(p in line.upper() for p in SKIP_PARTS)):
                fiche['nom'] = line.strip(' .')
                break
    except StopIteration:
        pass
    # Fallback : dernière ligne en CAPS significative
    if not fiche['nom']:
        for line in reversed(lines):
            if (len(line) >= 3 and line == line.upper()
                    and re.search(r'[A-ZÉÈÀÙÂÊÎÔÛÄËÏÖÜÇ]', line)
                    and line.strip(' .') not in SKIP_CAPS
                    and not any(p in line for p in SKIP_PARTS)
                    and not re.match(r'^[\d\s°.,:+×✓→]+$', line)
                    and not re.match(r'^\d', line)):
                fiche['nom'] = line.strip(' .')
                break

    # ── Positions des marqueurs ───────────────────────────────────────────────
    LM = {}
    for i, line in enumerate(lines):
        m = {
            'COEF': 'coef', 'Matériels et Équipements': 'mat',
            'Utilisations': 'util', 'NOTES DE SERVICES': 'ns',
            'APPARENCE NORMALE': 'an', 'APPARENCE ANORMALE': 'aa',
            'CONSERVATION': 'cons', 'Notes du Préparateur': 'np',
        }.get(line)
        if m and m not in LM:
            LM[m] = i

    def section(start_key, *end_keys):
        if start_key not in LM: return []
        s = LM[start_key] + 1
        ends = [LM[k] for k in end_keys if k in LM and LM[k] > s]
        return lines[s: min(ends) if ends else len(lines)]

    # Matériels
    fiche['materiels'] = [l for l in section('mat','util','ns','an','cons','np')
                          if l and not re.match(r'^ÉTAPE', l)]

    # Utilisations
    fiche['utilisations'] = '\n'.join(section('util','ns','an','cons','np')).strip() or None

    # Notes de services : le PDF mélange les ingrédients (layout 2 colonnes).
    # On ne garde rien ici — le champ sera à remplir manuellement.
    fiche['notes_services'] = None

    # ── Ingrédients depuis "Notes du Préparateur" ─────────────────────────────
    # Structure du PDF : [tous les noms] puis [toutes les quantités] (2 colonnes)
    # Parsé EN PREMIER pour construire known_ing_names utilisé dans apparence_normale.
    known_ing_names = set()
    if 'np' in LM:
        prep_lines = lines[LM['np'] + 1:]

        # Trouver la première ligne de quantité
        first_qty = next(
            (j for j, l in enumerate(prep_lines)
             if re.match(r'^[\d,.]+', l) or l in ('1 U', '1U')),
            None
        )

        if first_qty is not None and first_qty > 0:
            # Format [tous noms][toutes qtés] — cas habituel du PDF BOGOTA
            nom_lines = [l for l in prep_lines[:first_qty] if l and not l.isupper()]
            qty_lines  = [l for l in prep_lines[first_qty:]
                          if re.match(r'^[\d,.]+', l) or l in ('1 U', '1U')]
            ings = []
            for k, nom_l in enumerate(nom_lines):
                if k < len(qty_lines):
                    parts = qty_lines[k].split(None, 1)
                    ings.append({'nom': nom_l, 'qte': parts[0],
                                 'unite': parts[1].upper() if len(parts) > 1 else ''})
                else:
                    ings.append({'nom': nom_l, 'qte': '', 'unite': ''})
                known_ing_names.add(nom_l.lower())
            fiche['ingredients'] = ings
        else:
            # Fallback : paires [nom, qté] entrelacées
            ings, i = [], 0
            while i < len(prep_lines):
                nom_l = prep_lines[i]
                if i + 1 < len(prep_lines):
                    qty_l = prep_lines[i + 1]
                    if re.match(r'^[\d,.]+', qty_l) or qty_l in ('1 U', '1U'):
                        parts = qty_l.split(None, 1)
                        if nom_l and not nom_l.isupper():
                            ings.append({'nom': nom_l, 'qte': parts[0],
                                         'unite': parts[1].upper() if len(parts) > 1 else ''})
                            known_ing_names.add(nom_l.lower())
                        i += 2; continue
                i += 1
            fiche['ingredients'] = ings

    # Apparence normale :
    # La section NS→AN contient : [placeholder] [noms ingrédients] [quantités] [étapes?] [apparence]
    # → parcourir depuis la FIN et s'arrêter dès qu'on rencontre une instruction de cuisson.
    STEP_VERBS = re.compile(
        r'^(Ajouter|Remuer|Verser|Chauffer|Faire|Disposer|M[eé]langer|Finir|'
        r'Recouvrir|Laisser|Parsemer|Terminer|Lisser|puis|Retirer|Cuire|Porter|'
        r'Assaisonner|Incorporer|R[eé]server|Saler|Poivrer|Couvrir|D[eé]poser|'
        r'D[eé]couper|Garnir|Former|Rouler|Plonger|D[eé]glacer|Flamber|Monter)',
        re.IGNORECASE
    )
    if 'an' in LM:
        ref = LM.get('ns', LM.get('coef', 0)) + 1
        window = lines[ref:LM['an']]
        last_qty_idx = -1
        for idx, l in enumerate(window):
            if re.match(r'^[\d,.]', l):
                last_qty_idx = idx

        if last_qty_idx >= 0:
            # Remonter depuis la fin jusqu'à la première ligne "instruction"
            app_n = []
            for l in reversed(window[last_qty_idx + 1:]):
                if not l or l.isupper():
                    break
                if (l.endswith('.')             # phrase terminée par un point
                        or re.search(r'\(\d', l)  # quantité entre parenthèses : ex. "(120 g)"
                        or STEP_VERBS.match(l)):   # verbe de cuisson
                    break
                app_n.insert(0, l)
        else:
            # Aucune quantité : fallback avec filtre sur noms connus
            EXCL_HEADS = {'Matériels et Équipements', 'Utilisations', 'NOTES DE SERVICES'}
            app_n = []
            for l in reversed(window):
                if not l or l.isupper() or l.startswith('Voici'):
                    continue
                if (l.endswith('.') or re.search(r'\(\d', l) or STEP_VERBS.match(l)
                        or l.lower() in known_ing_names):
                    break
                app_n.insert(0, l)
        fiche['apparence_normale'] = '\n'.join(app_n).strip() or None

    # Apparence anormale
    if 'aa' in LM:
        start = LM.get('an', 0) + 1
        cands = [l for l in lines[start:LM['aa']]
                 if l and len(l) < 60 and not re.match(r'^\d', l)
                 and l.lower() not in known_ing_names]
        fiche['apparence_anormale'] = '\n'.join(cands).strip() or None

    # Conservation (sans les lignes en majuscules)
    fiche['conservation'] = '\n'.join(
        l for l in section('cons','np') if l and not l.isupper()
    ).strip() or None

    # ── Étapes ────────────────────────────────────────────────────────────────
    ETAPE_STOP = {'Matériels et Équipements','Utilisations','NOTES DE SERVICES',
                  'APPARENCE NORMALE','APPARENCE ANORMALE','CONSERVATION','Notes du Préparateur'}
    etapes, cur = [], []
    for line in lines[LM.get('coef', 0) + 1:]:
        if line in ETAPE_STOP: break
        em = re.match(r'ÉTAPES?\s*\d*\s*:\s*(.*)', line)
        if em:
            if cur: etapes.append(' '.join(cur))
            cur = [em.group(1).strip()] if em.group(1).strip() else []
        elif line and not (line.isupper() and len(line) > 3):
            cur.append(line)
    if cur: etapes.append(' '.join(cur))
    fiche['etapes'] = [e for e in etapes if e.strip()]

    return fiche if fiche['nom'] else None


# ── ROUTE : IMPORT PDF ────────────────────────────────────────────────────────
@app.post("/api/import/pdf")
async def import_pdf_route(file: UploadFile = File(...)):
    if not HAS_FITZ:
        raise HTTPException(400, "pymupdf non installé — lancez : pip install pymupdf")
    content = await file.read()
    try:
        doc = fitz.open(stream=content, filetype="pdf")
    except Exception as e:
        raise HTTPException(400, f"PDF invalide : {e}")

    recettes, vues = [], set()
    total = len(doc)
    for i, page in enumerate(doc):
        try:
            r = parse_fiche_pdf(page.get_text())
            if r and r.get('nom') and r['nom'] not in vues:
                r['page'] = i + 1
                recettes.append(r)
                vues.add(r['nom'])
        except Exception:
            pass
    doc.close()
    return {"recettes": recettes, "total_pages": total}


# ── ROUTE : SYNC HACCP ───────────────────────────────────────────────────────
HACCP_SHEETS = [
    ("PREPA PRIX NEW",              "prepa_cuisine"),
    ("SAUCE PRIX NEW",              "prepa_sauce"),
    ("PIZZA PRIX NEW",              "pizza"),
    ("FICHE TECH NOUVELLES PIZZAS", "pizza"),
    ("BOISSONS",                    "boisson"),
    ("SNACKS",                      "snack"),
]

def _haccp_recipes(ws):
    rows = list(ws.iter_rows(values_only=True))
    recipes, visited = [], set()
    for r_idx, row in enumerate(rows):
        for c_idx, cell in enumerate(row):
            if cell != "PRODUITS" or (r_idx, c_idx) in visited:
                continue
            visited.add((r_idx, c_idx))
            nom, rendement = None, None
            for rr in range(r_idx - 1, max(r_idx - 6, -1), -1):
                v = rows[rr][c_idx] if c_idx < len(rows[rr]) else None
                if isinstance(v, str) and len(v.strip()) > 1 and v.strip() not in (
                    "PRODUITS","QUANTITE","MESURE","PRIX","PREPARATION","RECETTE",
                    "G","KG","L","ML","g","kg","l","ml"
                ):
                    if nom is None:
                        nom = v.strip()
                if isinstance(v, (int, float)) and not isinstance(v, bool) and v and v > 0:
                    if rendement is None:
                        rendement = v
            if not nom:
                continue
            ings, cout_total = [], 0.0
            for ir in range(r_idx + 1, len(rows)):
                prod = rows[ir][c_idx] if c_idx < len(rows[ir]) else None
                if prod is None or prod == "" or prod == 0:
                    break
                if isinstance(prod, str) and prod.strip():
                    qte  = rows[ir][c_idx+1] if c_idx+1 < len(rows[ir]) else None
                    mes  = rows[ir][c_idx+2] if c_idx+2 < len(rows[ir]) else None
                    prix = rows[ir][c_idx+3] if c_idx+3 < len(rows[ir]) else None
                    if isinstance(prix, str): prix = None
                    if isinstance(prix, (int, float)) and prix:
                        cout_total += prix
                    q_str = str(int(qte)) if isinstance(qte, float) and qte == int(qte) else (str(qte) if qte is not None else "")
                    prix_ing = prix if isinstance(prix, (int, float)) and prix else None
                    ings.append({"nom": prod.strip(), "qte": q_str, "unite": str(mes).strip() if mes else "", "prix": prix_ing})
            if ings:
                recipes.append({
                    "nom": nom,
                    "cout": round(cout_total) if cout_total > 0 else None,
                    "quantite": int(rendement) if isinstance(rendement, (int, float)) else None,
                    "ingredients": ings,
                })
    return recipes


@app.post("/api/sync/haccp")
async def sync_haccp(file: UploadFile = File(...), request: Request = None, db: Session = Depends(get_db)):
    require_admin(request)
    if not HAS_OPENPYXL:
        raise HTTPException(400, "openpyxl non installé — lancez : pip install openpyxl")
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Fichier Excel invalide : {e}")

    created, updated, skipped = 0, 0, 0
    seen_noms = set()

    for sheet_name, categorie in HACCP_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        recipes = _haccp_recipes(wb[sheet_name])
        for rec in recipes:
            nom_key = rec["nom"].upper()
            if nom_key in seen_noms:
                continue
            seen_noms.add(nom_key)

            # Chercher par nom (insensible à la casse)
            existing = None
            all_f = db.query(Fiche).all()
            for f in all_f:
                if f.nom.upper() == nom_key:
                    existing = f
                    break

            now = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")

            if existing:
                # Mise à jour coût, quantité, catégorie, ingrédients
                changes = []
                if rec["cout"] and existing.cout != rec["cout"]:
                    changes.append({"champ": "Coût", "avant": existing.cout, "apres": rec["cout"]})
                    existing.cout = rec["cout"]
                if rec["quantite"] and existing.quantite != rec["quantite"]:
                    existing.quantite = rec["quantite"]
                if existing.categorie != categorie:
                    existing.categorie = categorie
                existing.updated_at = now
                # Remplacer ingrédients
                db.query(Ingredient).filter(Ingredient.fiche_id == existing.id).delete()
                for ing in rec["ingredients"]:
                    db.add(Ingredient(fiche_id=existing.id, nom=ing["nom"], qte=ing["qte"], unite=ing["unite"], prix=ing.get("prix")))
                if changes:
                    log_history(db, "modifier", existing.id, existing.nom, {"changes": changes, "source": "sync_haccp"})
                updated += 1
            else:
                fid = datetime.now().strftime("%Y%m%d%H%M%S") + "".join(random.choices(string.ascii_lowercase, k=4))
                new_f = Fiche(
                    id=fid, nom=rec["nom"], categorie=categorie,
                    quantite=rec["quantite"], unite="personnes",
                    cout=rec["cout"], created_at=now, updated_at=now
                )
                db.add(new_f)
                db.flush()
                for ing in rec["ingredients"]:
                    db.add(Ingredient(fiche_id=fid, nom=ing["nom"], qte=ing["qte"], unite=ing["unite"], prix=ing.get("prix")))
                log_history(db, "créer", fid, rec["nom"], {"categorie": categorie, "cout": rec["cout"], "nb_ingredients": len(rec["ingredients"]), "source": "sync_haccp"})
                created += 1

    # ── Feuille "Fiches" (format export application) ─────────────────────────
    if "Fiches" in wb.sheetnames:
        ws_f = wb["Fiches"]
        rows_f = list(ws_f.iter_rows(values_only=True))
        if rows_f:
            headers = [str(c).strip() if c is not None else "" for c in rows_f[0]]
            seen_fiches = set()
            for raw in rows_f[1:]:
                data_row = {h: v for h, v in zip(headers, raw)}
                nom = str(data_row.get("Nom") or "").strip()
                if not nom:
                    continue
                nom_key = nom.upper()
                if nom_key in seen_fiches:
                    continue
                seen_fiches.add(nom_key)

                row_id = str(data_row.get("ID") or "").strip()

                # Chercher existant : par ID d'abord, puis par nom
                existing = None
                if row_id:
                    existing = db.query(Fiche).filter(Fiche.id == row_id).first()
                if not existing:
                    for f2 in db.query(Fiche).all():
                        if f2.nom.upper() == nom_key:
                            existing = f2
                            break

                now = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")

                # Conservation
                c_val   = existing.conservation_val   if existing else None
                c_unite = existing.conservation_unite if existing else "jours"
                cons_raw = data_row.get("Conservation")
                if cons_raw:
                    parts = str(cons_raw).strip().split()
                    if len(parts) >= 2:
                        try:
                            c_val   = int(parts[0])
                            c_unite = " ".join(parts[1:])
                        except (ValueError, TypeError):
                            pass

                # Allergènes
                allerg_raw  = data_row.get("Allergènes")
                allerg_list = [a.strip() for a in str(allerg_raw).split(",") if a.strip()] if allerg_raw else []

                # Matériels (stockés \n-séparés)
                mat_raw = data_row.get("Matériels")
                mat_str = "\n".join(m.strip() for m in str(mat_raw).split(",") if m.strip()) if mat_raw else (existing.materiels if existing else None)

                def _v(key, fallback=None):
                    val = data_row.get(key)
                    return val if val is not None and val != "" else fallback

                if existing:
                    existing.nom = nom
                    if _v("Unité")             is not None: existing.unite              = _v("Unité")
                    if _v("Quantité")          is not None: existing.quantite           = _v("Quantité")
                    if _v("Coût matière (Ar)") is not None: existing.cout               = _v("Coût matière (Ar)")
                    if _v("Coefficient")       is not None: existing.coefficient        = float(_v("Coefficient"))
                    if _v("Préparation (min)") is not None: existing.prep               = int(_v("Préparation (min)"))
                    if _v("Cuisson (min)")     is not None: existing.cuisson            = int(_v("Cuisson (min)"))
                    if _v("Température (°C)")  is not None: existing.temperature        = _v("Température (°C)")
                    existing.conservation_val   = c_val
                    existing.conservation_unite = c_unite
                    if mat_str is not None:                  existing.materiels          = mat_str
                    if _v("Apparence normale"):   existing.apparence_normale  = _v("Apparence normale")
                    if _v("Apparence anormale"):  existing.apparence_anormale = _v("Apparence anormale")
                    if _v("Utilisations"):        existing.utilisations       = _v("Utilisations")
                    if _v("Notes de services"):   existing.notes_services     = _v("Notes de services")
                    if _v("Notes préparateur"):   existing.notes_preparateur  = _v("Notes préparateur")
                    existing.updated_at = now
                    if allerg_list:
                        db.query(FicheAllergene).filter(FicheAllergene.fiche_id == existing.id).delete()
                        for al in allerg_list:
                            db.add(FicheAllergene(fiche_id=existing.id, allergene=al))
                    updated += 1
                else:
                    fid   = row_id or gen_id()
                    coeff = float(_v("Coefficient") or 2.0)
                    prep_v   = _v("Préparation (min)")
                    cuiss_v  = _v("Cuisson (min)")
                    new_f = Fiche(
                        id=fid, nom=nom, categorie=None,
                        quantite=_v("Quantité"),
                        unite=_v("Unité") or "personnes",
                        cout=_v("Coût matière (Ar)"),
                        coefficient=coeff,
                        prep=int(prep_v) if prep_v is not None else None,
                        cuisson=int(cuiss_v) if cuiss_v is not None else None,
                        temperature=_v("Température (°C)"),
                        conservation_val=c_val, conservation_unite=c_unite,
                        materiels=mat_str,
                        apparence_normale=_v("Apparence normale"),
                        apparence_anormale=_v("Apparence anormale"),
                        utilisations=_v("Utilisations"),
                        notes_services=_v("Notes de services"),
                        notes_preparateur=_v("Notes préparateur"),
                        created_at=now, updated_at=now
                    )
                    db.add(new_f); db.flush()
                    for al in allerg_list:
                        db.add(FicheAllergene(fiche_id=fid, allergene=al))
                    log_history(db, "créer", fid, nom, {"source": "import_excel"})
                    created += 1

    db.commit()
    return {"created": created, "updated": updated, "skipped": skipped}


# ── STATIC + FRONTEND ────────────────────────────────────────────────────────
os.makedirs(STATIC_DIR, exist_ok=True)
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

@app.get("/")
def root():
    return FileResponse(os.path.join(STATIC_DIR, "index.html"))

# ── LAUNCH ────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    # PyInstaller --noconsole rend sys.stdout/stderr = None
    # Uvicorn plante car son formatter appelle .isatty() sur None
    # → on redirige vers un fichier log quand l'app est gelée sans console
    if _FROZEN and sys.stdout is None:
        _log_path = os.path.join(DATA_DIR, 'fiche-technique.log')
        _log_file = open(_log_path, 'w', encoding='utf-8', buffering=1)
        sys.stdout = _log_file
        sys.stderr = _log_file

    import socket, webbrowser, threading

    def port_libre(port):
        """True si personne n'écoute sur ce port."""
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            return s.connect_ex(('localhost', port)) != 0

    PORT = 8000

    if not port_libre(PORT):
        # Un serveur tourne déjà → ouvrir simplement le navigateur
        webbrowser.open(f"http://localhost:{PORT}")
    else:
        def open_browser():
            import time; time.sleep(1.5)
            webbrowser.open(f"http://localhost:{PORT}")
        threading.Thread(target=open_browser, daemon=True).start()
        # En .exe PyInstaller : passer l'objet app directement (pas de string "main:app")
        uvicorn.run(app, host="0.0.0.0", port=PORT, reload=False)
